"""
============================================
Bot Routes - Admin Command Handlers
============================================

This module contains all admin command handlers
for the Society Payment Tracker Bot.

Features:
- Table view (last 20 payments)
- Daily total
- Monthly total
- Member-specific history
- Excel export
- Data reset with confirmation
- Statistics
"""

import os
import logging
import tempfile
from datetime import datetime
from typing import Dict, Optional

from telegram import Update
from telegram.ext import ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from controllers.payment_controller import payment_controller

# Configure logging
logger = logging.getLogger(__name__)

# Store reset confirmations (user_id -> timestamp)
reset_confirmations: Dict[int, float] = {}
RESET_TIMEOUT = 60  # 60 seconds to confirm reset


def capitalize_first(text: str) -> str:
    """Capitalize the first letter of a string."""
    if not text:
        return text
    return text[0].upper() + text[1:].lower()


def format_short_date(dt: datetime) -> str:
    """Format date as YYYY-MM-DD."""
    return dt.strftime('%Y-%m-%d')


def format_full_datetime(dt: datetime) -> str:
    """Format date as YYYY-MM-DD HH:MM:SS."""
    return dt.strftime('%Y-%m-%d %H:%M:%S')


async def handle_table_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /table command - Show last 20 payments.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    
    try:
        payments = payment_controller.get_last_payments(20)
        
        if not payments:
            await context.bot.send_message(chat_id, '📭 No payment records found.')
            return
        
        # Build formatted table
        message = '📊 *Last 20 Payments*\n\n'
        message += '```\n'
        message += 'ID   | Member     | Amount    | Date\n'
        message += '-----+------------+-----------+------------\n'
        
        for payment in payments:
            pid = str(payment['id']).ljust(4)
            name = str(payment['member_name'])[:10].ljust(10)
            amount = f"Rs.{float(payment['amount']):.0f}".ljust(9)
            date = format_short_date(payment['payment_date'])
            
            message += f'{pid} | {name} | {amount} | {date}\n'
        
        message += '```'
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.info('📊 Table command executed')
        
    except Exception as e:
        logger.error(f'❌ Error in table command: {e}')
        await context.bot.send_message(chat_id, '❌ Error fetching payment records.')


async def handle_today_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /today command - Show today's total.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    
    try:
        result = payment_controller.get_today_total()
        
        message = (
            f"📅 *Today's Collection ({result['date']})*\n\n"
            f"💰 Total Amount: *Rs.{result['total']:.2f}*\n"
            f"📝 Number of Payments: *{result['count']}*"
        )
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.info('📅 Today command executed')
        
    except Exception as e:
        logger.error(f"❌ Error in today command: {e}")
        await context.bot.send_message(chat_id, "❌ Error fetching today's total.")


async def handle_month_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /month command - Show current month's total.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    
    try:
        result = payment_controller.get_month_total()
        
        message = (
            f"📆 *{result['month_name']} {result['year']} Collection*\n\n"
            f"💰 Total Amount: *Rs.{result['total']:.2f}*\n"
            f"📝 Number of Payments: *{result['count']}*"
        )
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.info('📆 Month command executed')
        
    except Exception as e:
        logger.error(f'❌ Error in month command: {e}')
        await context.bot.send_message(chat_id, "❌ Error fetching monthly total.")


async def handle_member_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /member command - Show member's payment history.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    
    # Check if member name is provided
    if not context.args:
        await context.bot.send_message(
            chat_id,
            '❌ Please provide a member name.\nUsage: /member <name>'
        )
        return
    
    member_name = context.args[0]
    
    try:
        result = payment_controller.get_member_payments(member_name)
        
        if not result['payments']:
            await context.bot.send_message(
                chat_id,
                f'❌ No payment records found for member: {member_name}'
            )
            return
        
        # Build member summary
        message = (
            f"👤 *Payment History: {capitalize_first(result['member_name'])}*\n\n"
            f"💰 Total Paid: *Rs.{result['total_amount']:.2f}*\n"
            f"📝 Total Payments: *{result['total_payments']}*\n\n"
            f"*Recent Payments:*\n"
        )
        
        # Show up to 10 recent payments
        recent_payments = result['payments'][:10]
        for payment in recent_payments:
            date = format_short_date(payment['payment_date'])
            message += f"• Rs.{float(payment['amount']):.2f} on {date}\n"
        
        if len(result['payments']) > 10:
            message += f"\n_... and {len(result['payments']) - 10} more payments_"
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.info(f'👤 Member command executed for: {member_name}')
        
    except Exception as e:
        logger.error(f'❌ Error in member command: {e}')
        await context.bot.send_message(chat_id, '❌ Error fetching member records.')


async def handle_export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /export command - Export all data to Excel.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    filepath = None
    
    try:
        # Notify user that export is in progress
        await context.bot.send_message(chat_id, '📤 Generating Excel export...')
        
        payments = payment_controller.get_all_payments()
        
        if not payments:
            await context.bot.send_message(chat_id, '📭 No records to export.')
            return
        
        logger.info(f'📤 Preparing export for {len(payments)} records...')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payments'
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Member Name', 'Amount (Rs.)', 'Recorded By (User ID)', 'Payment Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        total_amount = 0.0
        for row_num, payment in enumerate(payments, 2):
            amount = float(payment['amount'])
            total_amount += amount
            
            ws.cell(row=row_num, column=1, value=payment['id']).border = thin_border
            ws.cell(row=row_num, column=2, value=capitalize_first(payment['member_name'])).border = thin_border
            ws.cell(row=row_num, column=3, value=f"{amount:.2f}").border = thin_border
            ws.cell(row=row_num, column=4, value=str(payment['recorded_by'])).border = thin_border
            ws.cell(row=row_num, column=5, value=format_full_datetime(payment['payment_date'])).border = thin_border
        
        # Summary row
        summary_row = len(payments) + 3
        ws.cell(row=summary_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=f"{total_amount:.2f}").font = Font(bold=True)
        
        # Set column widths
        column_widths = [8, 20, 15, 22, 22]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f'society_payments_{timestamp}.xlsx'
        filepath = os.path.join(tempfile.gettempdir(), filename)
        
        logger.info(f'📤 Writing Excel file to: {filepath}')
        
        # Save workbook
        wb.save(filepath)
        
        # Verify file exists
        if not os.path.exists(filepath):
            raise Exception('Excel file was not created')
        
        file_size = os.path.getsize(filepath)
        logger.info(f'📤 Excel file created, size: {file_size} bytes')
        
        # Send file to Telegram
        with open(filepath, 'rb') as file:
            await context.bot.send_document(
                chat_id,
                document=file,
                filename=filename,
                caption=(
                    f"📊 Society Payments Export\n"
                    f"📝 Total Records: {len(payments)}\n"
                    f"💰 Total Amount: Rs.{total_amount:.2f}"
                )
            )
        
        logger.info(f'✅ Export completed: {len(payments)} records sent')
        
    except Exception as e:
        logger.error(f'❌ Error in export command: {e}')
        await context.bot.send_message(chat_id, f'❌ Error generating export file: {e}')
    
    finally:
        # Clean up temporary file
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
                logger.info('🗑️ Temporary file deleted')
            except Exception as e:
                logger.warning(f'⚠️ Could not delete temp file: {e}')


async def handle_reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /reset command - Ask for confirmation.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    
    try:
        # Get current stats before reset
        stats = payment_controller.get_payment_stats()
        
        if stats['total_payments'] == 0:
            await context.bot.send_message(chat_id, '📭 No records to delete.')
            return
        
        # Store reset confirmation request with timestamp
        reset_confirmations[user_id] = datetime.now().timestamp()
        
        message = (
            "⚠️ *WARNING: Reset Confirmation Required*\n\n"
            "You are about to delete ALL payment records:\n"
            f"• Total Records: *{stats['total_payments']}*\n"
            f"• Total Amount: *Rs.{stats['total_amount']:.2f}*\n"
            f"• Unique Members: *{stats['unique_members']}*\n\n"
            "This action *CANNOT BE UNDONE*.\n\n"
            "To confirm, type: /confirm_reset\n"
            "This confirmation will expire in 60 seconds."
        )
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.warning(f'⚠️ Reset requested by user {user_id}')
        
    except Exception as e:
        logger.error(f'❌ Error in reset command: {e}')
        await context.bot.send_message(chat_id, '❌ Error processing reset request.')


async def handle_confirm_reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /confirm_reset command - Execute the reset.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    
    try:
        # Check if this user has a valid reset confirmation pending
        confirmation_time = reset_confirmations.get(user_id)
        now = datetime.now().timestamp()
        
        if not confirmation_time or (now - confirmation_time) >= RESET_TIMEOUT:
            await context.bot.send_message(
                chat_id,
                '❌ No valid reset request found or it has expired.\nPlease use /reset first.'
            )
            # Clean up expired confirmation
            reset_confirmations.pop(user_id, None)
            return
        
        # Remove the confirmation
        reset_confirmations.pop(user_id, None)
        
        # Execute reset
        result = payment_controller.reset_all_payments()
        
        message = (
            "🗑️ *Reset Complete*\n\n"
            f"Successfully deleted *{result['deleted_count']}* payment records.\n"
            "The database is now empty."
        )
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.warning(f"🗑️ Reset executed by admin {user_id}: {result['deleted_count']} records deleted")
        
    except Exception as e:
        logger.error(f'❌ Error executing reset: {e}')
        await context.bot.send_message(chat_id, '❌ Error executing reset.')


async def handle_stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /stats command - Show payment statistics.
    
    Args:
        update: Telegram update object
        context: Callback context
    """
    chat_id = update.effective_chat.id
    
    try:
        stats = payment_controller.get_payment_stats()
        
        if stats['total_payments'] == 0:
            await context.bot.send_message(chat_id, '📭 No payment records found.')
            return
        
        message = (
            "📈 *Payment Statistics*\n\n"
            f"📝 Total Payments: *{stats['total_payments']}*\n"
            f"👥 Unique Members: *{stats['unique_members']}*\n\n"
            "💰 *Amount Summary:*\n"
            f"• Total: *Rs.{stats['total_amount']:.2f}*\n"
            f"• Average: *Rs.{stats['average_amount']:.2f}*\n"
            f"• Highest: *Rs.{stats['max_amount']:.2f}*\n"
            f"• Lowest: *Rs.{stats['min_amount']:.2f}*"
        )
        
        await context.bot.send_message(chat_id, message, parse_mode='Markdown')
        logger.info('📈 Stats command executed')
        
    except Exception as e:
        logger.error(f'❌ Error in stats command: {e}')
        await context.bot.send_message(chat_id, '❌ Error fetching statistics.')
